"""
MS Invoice Tool - Streamlit Application
Converts Excel files to MS Invoice format
"""

import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from msinvoice_processor import (
    process_ms_invoice_file,
    standardize_input_columns,
    validate_input_file,
    OUTPUT_HEADER,
)
from msinvoice_srcl import build_kuwait_exchange_lookup, create_ms_srcl_file

st.set_page_config(page_title="MS Invoice Tool", layout="wide")

st.markdown(
    """
    <style>
    [data-testid="stToolbar"],
    [data-testid="stHeaderActionElements"],
    .stAppDeployButton,
    [data-testid="stDecoration"],
    [data-testid="stStatusWidget"],
    [data-testid="stFloatingActionButton"] {
        display: none !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("MS Invoice Tool")


def normalize_date_key(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text_value = str(value).strip()
    if not text_value:
        return ""
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return text_value



# File upload
uploaded_file = st.file_uploader(
    "Upload your MS Invoice Excel file", 
    type=["xlsx", "xls"],
    key="ms_invoice_upload"
)

if uploaded_file:
    try:
        # Read the Excel file
        df = standardize_input_columns(pd.read_excel(uploaded_file))
        
        st.success(f"✅ File loaded: {uploaded_file.name} ({len(df)} rows)")
        
        # Validate input
        is_valid, validation_errors = validate_input_file(df)
        
        if not is_valid:
            st.error("❌ Validation failed:")
            for error in validation_errors:
                st.error(f"  • {error}")
            st.stop()
        
        # Process the file
        with st.spinner("Processing file..."):
            output_df, processing_errors = process_ms_invoice_file(df)
        
        # Display results
        if processing_errors:
            st.warning("⚠️ Some rows had processing issues:")
            for error in processing_errors[:10]:  # Show first 10 errors
                st.warning(f"  • {error}")
            if len(processing_errors) > 10:
                st.warning(f"  ... and {len(processing_errors) - 10} more")
        
        invoice_types = output_df.get("_Invoice Type", pd.Series("", index=output_df.index)).fillna("").astype(str).str.strip().str.lower()
        positive_df = output_df[invoice_types == "debit invoice"].copy()
        negative_df = output_df[invoice_types == "credit invoice"].copy()
        preview_df = positive_df[OUTPUT_HEADER].copy()

        kuwait_negative_df = negative_df[negative_df["Document Location"] == "WT000"].copy()
        kuwait_rate_lookup, kuwait_negative_dates, kuwait_ambiguous_dates = build_kuwait_exchange_lookup(df)
        ambiguous_negative_dates = [date_key for date_key in kuwait_negative_dates if date_key in kuwait_ambiguous_dates]
        unresolved_kuwait_dates = sorted(
            {
                str(date_key)
                for date_key in kuwait_negative_df.get("_Source Invoice Date", pd.Series(dtype=object)).map(normalize_date_key)
                if date_key and date_key not in kuwait_rate_lookup
            }
        )
        kuwait_manual_rate = None
        if not kuwait_negative_df.empty:
            kuwait_rate_input = st.text_input(
                "Kuwait SRCL exchange rate override",
                help="Used only for Kuwait negative rows when an exact same-date Kuwait positive exchange rate is not available.",
            ).strip()
            if kuwait_rate_input:
                try:
                    kuwait_manual_rate = float(kuwait_rate_input)
                    if kuwait_manual_rate <= 0:
                        st.error("Kuwait SRCL exchange rate must be greater than 0.")
                        kuwait_manual_rate = None
                except ValueError:
                    st.error("Kuwait SRCL exchange rate must be a valid number.")
                    kuwait_manual_rate = None
        

        # Create Excel file for download
        output_buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "MS INVOICE"
        
        # Write header
        ws.append(OUTPUT_HEADER)
        
        # Write data
        for _, row in preview_df.iterrows():
            row_list = [row.get(col, "") for col in OUTPUT_HEADER]
            ws.append(row_list)

        date_columns = {
            "Invoice Date",
            "Delivery Date",
            "Billing Cycle Start Date",
            "Billing Cycle End Date",
        }
        for col_idx, col_name in enumerate(OUTPUT_HEADER, start=1):
            if col_name in date_columns:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "dd/mm/yyyy"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        # Download button
        st.download_button(
            label="⬇️ Download MS Invoice (Excel)",
            data=output_buffer.getvalue(),
            file_name="ms_invoice_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if not negative_df.empty:
            if ambiguous_negative_dates and kuwait_manual_rate is None:
                st.error(
                    "SRCL was not generated because Kuwait has multiple positive exchange rates on these dates: "
                    + ", ".join(ambiguous_negative_dates)
                    + ". Enter a Kuwait SRCL exchange rate override to continue."
                )
            elif unresolved_kuwait_dates and kuwait_manual_rate is None:
                st.error(
                    "SRCL was not generated because Kuwait negative rows were found without a same-date Kuwait "
                    "positive exchange rate on these dates: "
                    + ", ".join(unresolved_kuwait_dates)
                    + ". Enter a Kuwait SRCL exchange rate override to continue."
                )
            else:
                srcl_buffer = create_ms_srcl_file(
                    negative_df,
                    kuwait_rate_lookup=kuwait_rate_lookup,
                    kuwait_manual_rate=kuwait_manual_rate,
                )
                st.download_button(
                    label="⬇️ Download SRCL File",
                    data=srcl_buffer.getvalue(),
                    file_name="ms_srcl_file.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        
    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        st.error("Please ensure your file is in the correct Excel format.")
