import io

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from utils import validate_customer_code, build_cloud_invoice_df, map_invoice_numbers, create_srcl_file
from io import BytesIO
st.title("Cloud Invoice Tool")
st.markdown(
        """
        <div style="
            padding: 18px 20px;
            background: linear-gradient(90deg, #fff3cd, #ffeeba);
            border: 2px solid #ffcc00;
            border-radius: 10px;
            font-weight: 700;
            color: #7a5a00;
            font-size: 16px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            margin-bottom: 12px;
        ">
            <span style="font-size: 18px;">🚨 IMPORTANT:</span>
            <span style="margin-left: 8px;">
            Please make sure to <b>open the CB file</b>, click <b>Convert</b>, then <b>upload the converted file here</b> and use this tool; otherwise <b>you will have missing invoices</b>.
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )
st.markdown(
        """
        <div style="
            padding: 14px 16px;
            background: #fff;
            border: 1px dashed #ffcc00;
            border-radius: 10px;
            color: #4a4a4a;
            font-size: 15px;
            margin-bottom: 8px;
        ">
        <b>Follow these steps before uploading:</b>
        <ol style="margin-top: 6px;">
            <li>Open the <b>CB file</b>.</li>
            <li>Click <b>Convert</b> to generate the latest output.</li>
            <li>Upload the <b>converted file</b> here.</li>
        </ol>
        </div>
        """,
        unsafe_allow_html=True,
    )
confirmed = st.checkbox("I confirm I opened the CB file and clicked Convert ✅", key="cloud_cb_confirm")
if not confirmed:
        st.warning("Please confirm the IMPORTANT notice steps above to proceed.")
        st.stop()
uploaded_file = st.file_uploader("Upload your CSV file", type=["csv", "xlsx"], key="cloud_invoice_upload")
if uploaded_file:
        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith(".xlsx"):
            df = pd.read_excel(uploaded_file, engine="openpyxl")
        else:
            st.error("Unsupported file format. Please upload a CSV or Excel file.")
            st.stop()
            # --- Validate Customer Code before proceeding ---
        validate_customer_code(df, "Cloud Invoice File")

        # Process invoice data
        final_df = build_cloud_invoice_df(df)
        final_df = map_invoice_numbers(final_df)
        sorted_df = final_df.sort_values(by=final_df.columns.tolist()).reset_index(drop=True)
        def highlight_row(row):
            end_user = str(row.get("End User", "")).strip()
            return not (" ; " in end_user)
        sorted_df["_highlight_end_user"] = sorted_df.apply(highlight_row, axis=1)
        # Create unique version rows based on Combined (D)
        unique_rows = sorted_df[["Invoice No.","LPO Number", "End User"]].copy()
        unique_rows["Combined (D)"] = (
            unique_rows["Invoice No."].astype(str) +
            unique_rows["LPO Number"].astype(str) +
            unique_rows["End User"].astype(str)
        )
        unique_rows = unique_rows.drop_duplicates(subset=["Combined (D)"]).reset_index(drop=True)
        # Versioning logic
        unique_rows["Version1 (E)"] = (unique_rows["Invoice No."].ne(unique_rows["Invoice No."].shift()).astype(int))
        v2 = []
        for i, v1 in enumerate(unique_rows["Version1 (E)"]):
            if v1 == 1:
                v2.append(1)
            else:
                prev_v2 = v2[-1]
                v2.append(prev_v2 + 1)
        unique_rows["Version2 (F)"] = v2
        unique_rows["Version3 (G)"] = unique_rows.apply(lambda row: f'-{row["Version2 (F)"]}', axis=1)
        unique_rows["Version4 (H)"] = unique_rows.apply(lambda row: f'{row["Invoice No."]}-{row["Version2 (F)"]}', axis=1)
        # --- MAP Version 4 back to main DataFrame ---
        version_map = dict(zip(unique_rows["Combined (D)"], unique_rows["Version4 (H)"]))
        sorted_df["Combined (D)"] = (
            sorted_df["Invoice No."].astype(str) +
            sorted_df["LPO Number"].astype(str) +
            sorted_df["End User"].astype(str)
        )
        sorted_df["Versioned Invoice No."] = sorted_df["Combined (D)"].map(version_map)
        cols = list(sorted_df.columns)
        cols.append(cols.pop(cols.index("Versioned Invoice No.")))
        sorted_df = sorted_df[cols]
        sorted_df = sorted_df.drop(columns=["Combined (D)"])
        # === ADD HIGHLIGHT FLAG HERE ===
        #sorted_df["_highlight_end_user"] = sorted_df["End User"].astype(str).str.strip() == ""
        # Display metrics
        pos_df = sorted_df[sorted_df["Gross Value"].astype(float) >= 0]
        neg_df = sorted_df[sorted_df["Gross Value"].astype(float) < 0]
        st.success(f"{len(pos_df)} positive, {len(neg_df)} negative, total: {len(sorted_df)}")
        c1, c2, c3 = st.columns(3)
        c1.metric("✅ Positive invoices", len(pos_df))
        c2.metric("❌ Negative invoices", len(neg_df))
        c3.metric("🧮 Total invoices", len(sorted_df))
       
        
        # Create Excel workbook with formulas
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        df_to_write = pos_df.copy()
        
        # Remove highlight flag column before writing
        if "_highlight_end_user" in df_to_write.columns:
            df_to_write = df_to_write.drop(columns=["_highlight_end_user"])
        
        # Get index of 'End User' column (1-based for Excel)
        try:
            end_user_col_idx = df_to_write.columns.get_loc("End User") + 1
        except:
            end_user_col_idx = None
            
        try:
            item_code_col_idx = df_to_write.columns.get_loc("ITEM Code") + 1
        except:
            item_code_col_idx = None    
        
        # Create workbook and write rows
        wb = Workbook()
        ws_invoice = wb.active
        ws_invoice.title = "CLOUD INVOICE"
        
        for r_idx, row in enumerate(dataframe_to_rows(df_to_write, index=False, header=True), start=1):
            ws_invoice.append(row)
            
            # Skip header row
            if r_idx == 1:
                continue
            # Highlight End User
            if end_user_col_idx is not None:
                highlight = sorted_df.iloc[r_idx - 2].get("_highlight_end_user", False)
                if highlight:
                    col_letter = get_column_letter(end_user_col_idx)
                    ws_invoice[f"{col_letter}{r_idx}"].fill = red_fill
            
            # Highlight ITEM Code if empty
            if item_code_col_idx is not None:
                item_code_val = sorted_df.iloc[r_idx - 2].get("ITEM Code", "")
                if not item_code_val or str(item_code_val).strip().lower() in ["", "nan", "none"]:
                    col_letter = get_column_letter(item_code_col_idx)
                    ws_invoice[f"{col_letter}{r_idx}"].fill = red_fill
        
        # Create VERSIONS sheet with formulas
        ws_versions = wb.create_sheet(title="VERSIONS")
        headers = ["Invoice",  "LPO", "End User", "Combined (D)", "Version1 (E)", "Version2 (F)", "Version3 (G)", "Version4 (H)"]
        ws_versions.append(headers)
        for i, row in enumerate(unique_rows.itertuples(index=False, name=None), start=2):
            invoice, lpo, end_user, combined_d = row[:4]
            ws_versions.cell(row=i, column=1, value=invoice)
            ws_versions.cell(row=i, column=2, value=lpo)
            ws_versions.cell(row=i, column=3, value=end_user)
            ws_versions.cell(row=i, column=4, value=combined_d)
            ws_versions.cell(row=i, column=5, value=f'=IF(A{i}=A{i-1},"",1)')
            ws_versions.cell(row=i, column=6, value=f'=IFERROR(IF(E{i}="",E{i-1}+1,""),F{i-1}+1)')
            ws_versions.cell(row=i, column=7, value=f'="-"&E{i}&F{i}')
            ws_versions.cell(row=i, column=8, value=f'=A{i}&G{i}')
        
        # Save to buffer
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        st.download_button(
            label="⬇️ Download Cloud Invoice",
            data=output_buffer.getvalue(),
            file_name="cloud_invoice.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        neg_buffer = io.BytesIO()
        wb_neg = Workbook()
        ws_neg = wb_neg.active
        ws_neg.title = "NEGATIVE INVOICES"
        for row in dataframe_to_rows(neg_df, index=False, header=True):
            ws_neg.append(row)
        wb_neg.save(neg_buffer)
        neg_buffer.seek(0)
        st.download_button(
            label="⬇️ Download Negative Invoices",
            data=neg_buffer.getvalue(),
            file_name="negative_invoices.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        srcl_buffer = create_srcl_file(neg_df)  # only negative invoices
        
        st.download_button(
            label="⬇️ Download SRCL File",
            data=srcl_buffer.getvalue(),
            file_name="srcl_file.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )